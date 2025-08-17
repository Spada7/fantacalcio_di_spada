import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Font, Alignment

def copia_foglio_con_stile(foglio_sorgente, workbook_destinazione, nome_foglio_nuovo):
    foglio_nuovo = workbook_destinazione.create_sheet(title=nome_foglio_nuovo)

    grigio_chiaro = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
    verde_chiaro = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")

    for row in foglio_sorgente.iter_rows():
        for cell in row:
            nuova_cella = foglio_nuovo.cell(row=cell.row, column=cell.column, value=cell.value)

            # Riempimento originale
            if cell.fill and isinstance(cell.fill, PatternFill):
                nuova_cella.fill = PatternFill(
                    fill_type=cell.fill.fill_type,
                    start_color=cell.fill.start_color.rgb,
                    end_color=cell.fill.end_color.rgb
                )

            # Bordo
            if cell.border:
                nuova_cella.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )

            # Font
            if cell.font:
                nuova_cella.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    underline=cell.font.underline,
                    color=cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None
                )

            # Allineamento
            if cell.alignment:
                nuova_cella.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=cell.alignment.wrap_text
                )

            # Evidenziazione personalizzata
            valore = cell.value
            colore_font = cell.font.color.rgb if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb') else ""

            if isinstance(valore, (int, float)) and colore_font:
                if "00FF00" in colore_font.upper():  # verde
                    nuova_cella.fill = verde_chiaro

            elif valore is None or (isinstance(valore, str) and valore.strip() == ""):
                nuova_cella.fill = grigio_chiaro

    # Larghezza colonne
    for col in foglio_sorgente.column_dimensions:
        foglio_nuovo.column_dimensions[col].width = foglio_sorgente.column_dimensions[col].width

    # Celle unite
    for merged_range in foglio_sorgente.merged_cells.ranges:
        foglio_nuovo.merge_cells(str(merged_range))

def main():
    file_quotazioni = 'Quotazioni_Fantacalcio_Stagione_2025_26.xlsx'
    file_statistiche = 'Statistiche_Fantacalcio_Stagione_2024_25.xlsx'
    griglia_porta = 'Griglia_Portieri_Fantacalcio_Stagione_2025-26.xlsx'
    output_file = 'Output_Fantacalcio_Classico.xlsx'

    df_q = pd.read_excel(file_quotazioni, header=1)
    df_q = df_q[['Nome', 'Squadra', 'R', 'Qt.A']].copy()
    df_q.rename(columns={'R': 'Ruolo', 'Qt.A': 'Quotazione'}, inplace=True)

    df_s = pd.read_excel(file_statistiche, header=1)
    stats_columns = [col for col in df_s.columns if col not in ['Id', 'R', 'Rm', 'Squadra']]
    df_s = df_s[stats_columns]

    merged_df = df_q.merge(df_s, on='Nome', how='left')
    merged_df.sort_values(by=['Squadra', 'Ruolo', 'Quotazione'], ascending=[True, True, False], inplace=True)

    portieri = merged_df[merged_df['Ruolo'] == 'P']
    difensori = merged_df[merged_df['Ruolo'] == 'D']
    centrocampisti = merged_df[merged_df['Ruolo'] == 'C']
    attaccanti = merged_df[merged_df['Ruolo'] == 'A']

    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        portieri.to_excel(writer, sheet_name='Portieri', index=False)
        difensori.to_excel(writer, sheet_name='Difensori', index=False)
        centrocampisti.to_excel(writer, sheet_name='Centrocampisti', index=False)
        attaccanti.to_excel(writer, sheet_name='Attaccanti', index=False)

    wb_source = load_workbook(griglia_porta)
    wb_target = load_workbook(output_file)

    ws_source = wb_source.active
    copia_foglio_con_stile(ws_source, wb_target, "Griglia Portieri")

    wb_target.save(output_file)
    print(f'✅ File Excel generato con successo: {output_file}')

if __name__ == '__main__':
    main()
